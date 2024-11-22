import os
import json
import random
from datetime import datetime
from flask import Blueprint
from flask import request, session
from flask import render_template, url_for, redirect, jsonify, send_file
from flask_login import login_required
from flask_login import current_user
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fonts import Font
from openpyxl.formatting.rule import Rule
from app.utils.db import execute_query


bp = Blueprint('exam', __name__)


def check_authority(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if current_user.is_authenticated:
            query = r"SELECT 1 FROM admin WHERE id=%s"
            if execute_query(query, (current_user.id)):
                return func(*args, **kwargs)
        return redirect("/")
    return wrapper



@bp.route("/lists")
@login_required
def exam_lists():
    query = r"SELECT s.id, s.name, COUNT(q.id) AS count FROM subjects s LEFT JOIN quizzes q ON q.s_id=s.id GROUP BY s.id, s.name"
    subjects = rows if (rows:=execute_query(query)) else []
    if not subjects:
        return redirect("/")
    # params = []
    # num_of_rows = rows[0]["num"] if (rows:=execute_query(r"SELECT COUNT(*) FROM quizzes"+condition, tuple(params)))
    return render_template("exam/lists.html", subjects=subjects)


@bp.route("/setting/<idx>", methods=['GET', 'POST'])
@login_required
def exam_setting(idx: str):
    query = r"SELECT s.id, s.name, COUNT(q.id) AS count FROM subjects s LEFT JOIN quizzes q ON q.s_id=s.id WHERE s.id=%s GROUP BY s.id, s.name"
    subject = rows[0] if (rows:=execute_query(query, (idx,))) else []
    if not subject:
        redirect("/")
    return render_template("exam/setting.html", subject=subject)


@bp.route("/start", methods=["GET", 'POST'])
@login_required
def exam_start():
    json_data = request.form.get("settingData")
    data = json.loads(json_data)
    
    if not json_data or not data['s_id'].isdigit() or not data['count'].isdigit():
        return redirect("/")
    
    query = r"SELECT x.id, y.name s, x.question q, x.answer a, "
    query += r"x.opt1 o1, x.opt2 o2, x.opt3 o3, x.comment c "
    query += r"FROM quizzes x INNER JOIN subjects y ON x.s_id=%s AND x.s_id=y.id ORDER BY "
    query += r"x.id DESC LIMIT %s" if (data['exam_type'] == 1) else r"RAND() LIMIT %s"
    rows = execute_query(query, (int(data['s_id']), int(data['count'])))
    
    if not rows:
        return redirect("/")

    workbook = []
    answers = []
    options = {1: "a", 2: "o1", 3: "o2", 4: "o3"}

    for row in rows:
        seq = [1, 2, 3, 4]
        random.shuffle(seq)
        answers.append((row['id'], str(seq.index(1) + 1)))
        workbook.append({
            's': row['s'],
            "q": row["q"],
            "o1": row[options[seq[0]]],
            "o2": row[options[seq[1]]],
            "o3": row[options[seq[2]]],
            "o4": row[options[seq[3]]],
            "c": row["c"]
        })
    
    session.pop('workbook', None)
    session.pop('answers', None)
    session['workbook'] = workbook
    session['answers'] = answers

    return render_template('exam/start.html', workbook=workbook)


@bp.route("/submit", methods=['POST'])
@login_required
def exam_submit():
    json_data = request.form.get("userChoice")
    if not json_data:
        return redirect("/")
    
    answers = session["answers"]
    marking = json.loads(json_data)

    if len(answers) != len(marking):
        return redirect("/")

    wrong_list = [a[0] for a, b in zip(answers, marking) if a[1] != b]
    session.pop('marking', None)
    session["marking"] = marking

    return render_template("exam/result.html", num_of_quizzes=len(answers), correct_answer=(len(answers) - len(wrong_list)))



@bp.route("/download_result", methods=['GET'])
@login_required
def exam_download_result():
    w = session.get("workbook")
    a = session.get("answers")
    m = session.get("marking")

    if not (w and a and m):
        return redirect("/")
    
    data = [
        [
            quiz['s'],
            f"{quiz['q']}\n\n① {quiz['o1']}\n② {quiz['o2']}\n③ {quiz['o3']}\n④ {quiz['o4']}",
            "" if m[i] == "0" else m[i],
            a[i][1],
            quiz["c"]
        ]
        for i, quiz in enumerate(w)
    ]

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "시험결과"
    sheet.append(["과목", "문제", "마킹", "정답", "해설"])
    for row in data:
        sheet.append(row)
    
    # 서식 지정
    ## 행 및 열 크기 지정
    sheet.row_dimensions[1].hegiht = 30
    column_widths = {"A": 30, "B": 100, "C": 10, "D": 10, "E": 50}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    ## 가운데 정렬
    for area in ["A2:A100", "C2:D100"]:
        for row in sheet[area]:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # "문제" 열 줄바꿈 서식
    for row in sheet["B2:B100"]:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # 행 제목 서식
    for row in sheet["A1:E1"]:
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 조건부 서식
    fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
    rule = Rule(type="expression", formula=["$C2<>$D2"], dxf=DifferentialStyle(fill=fill))
    sheet.conditional_formatting.add("C2:C200", rule)

    # 파일 이름 설정
    num = (str(datetime.now().timestamp() * 1000000).replace(".",""))[-6:]
    file_path = os.path.join(os.getcwd(), "app", "tmp", f"quiz-{num}.xlsx")

    workbook.save(filename=file_path)
    workbook.close()

    return send_file(file_path, as_attachment=True)











# @bp.route("/lists")
# @login_required
# def exam_lists():
#     subjects = execute_query(r"SELECT id, name FROM subjects")
#     if subjects is False:
#         return redirect("/")
    
#     page = request.args.get("page", 1, type=int)
#     keyword = request.args.get("keyword", "", type=str)
#     sort_by = request.args.get("sortBy", "DESC", type=str)
#     search_by = request.args.get("searchBy", 0, type=int)

#     # 페이지네이션
#     total_pages = 0 # 전체 게시글 수
#     start_page = 0  # 표시할 첫번째 페이지
#     end_page = 0    # 표시할 마지막 페이지
#     per_page = 10   # 한 페이지에 표시할 게시글 수
#     offset = (page - 1) * per_page # 페이지 오프셋

#     # count_query = f"SELECT COUNT(*) num FROM exam"
#     # query = r"SELECT id, (SELECT name FROM users WHERE id=u_id) AS writer, title, content, DATE_FORMAT(created_at, '%%Y-%%m-%%d') AS created_at FROM lecture "
#     # if search_by == 0:
#     #     condition = f" WHERE title LIKE %s OR content LIKE %s ORDER BY 1 {sort_by}"
#     #     params = (f"%{keyword}%", f"%{keyword}%")
#     # elif search_by == 1:
#     #     condition = f" WHERE title LIKE %s ORDER BY 1 {sort_by}"
#     #     params = (f"%{keyword}%",)
#     # elif search_by == 2:
#     #     condition = f" WHERE content LIKE %s ORDER BY 1 {sort_by}"
#     #     params = (f"%{keyword}%",)
#     # limit = f" LIMIT {per_page} OFFSET {offset}"
    
#     # total_exams = rows[0]["num"] if (rows:=execute_query(count_query+condition, params)) else 0
#     # total_pages = (total_exams + per_page - 1) // per_page
#     # exams = rows if (rows:=execute_query(query+condition+limit, params)) else []

#     # start_page = max(page - 2, 1)
#     # end_page = min(page + 2, total_pages)

#     # if (end_page - start_page) < 4:
#     #     if start_page == 1:
#     #         end_page = min(start_page + 4, total_pages)
#     #     elif end_page == total_pages:
#     #         start_page = max(end_page - 4, 1)

#     data = {
#         "subjects": subjects,
#         "exams": [],
#         "page": page,
#         "total_pages": total_pages,
#         "start_page": start_page,
#         "end_page": end_page
#     }
#     return render_template("exam/lists.html", **data)


# @bp.route("/create")
# @login_required
# @check_authority
# def exam_create():
#     return render_template("exam/create.html")


# @bp.route("/modify")
# @login_required
# @check_authority
# def exam_modify():
#     return render_template("exam/modify.html")


# @bp.route("/view")
# @login_required
# @check_authority
# def exam_view():
#     return render_template("exam/view.html")